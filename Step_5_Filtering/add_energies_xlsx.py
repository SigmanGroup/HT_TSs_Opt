#!/usr/bin/env python

import os
from openpyxl import load_workbook

# Function to read the Excel file and store data in a dictionary
def read_excel(excel_filename):
    data = {}
    workbook = load_workbook(excel_filename)
    sheet = workbook.active
    for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip the header row
        data[row[0]] = float(row[1])
    return data

# Function to replace string in XYZ file
def replace_string_in_file(filename, old_string, new_string):
    with open(filename, 'r') as file:
        content = file.read()

    content = content.replace(old_string, new_string)

    with open(filename, 'w') as file:
        file.write(content)

# Main function
def main():
    excel_filename = "xTB_E.xlsx"

    # Read data from Excel file
    data = read_excel(excel_filename)

    # Get list of XYZ files in the directory
    xyz_files = [f for f in os.listdir() if f.endswith(".xyz")]

    if len(xyz_files) == 0:
        print("No XYZ files found in the directory.")
        return

    # Process each XYZ file
    for xyz_filename in xyz_files:
        # Extract base filename without extension
        base_filename = os.path.splitext(xyz_filename)[0]

        # Check if corresponding string exists in XYZ file
        string_to_replace = f"{base_filename}.log"
        with open(xyz_filename, 'r') as file:
            xyz_content = file.read()

        if string_to_replace in xyz_content:
            # Replace string with filename, tab, 'Eopt', and corresponding float from Excel
            new_float = data.get(f"{base_filename}", None)
            if new_float is not None:
                replacement_string = f"{base_filename}\tEopt\t{new_float}"
                replace_string_in_file(xyz_filename, string_to_replace, replacement_string)
                print(f"String replaced successfully in {xyz_filename}.")
            else:
                print(f"Corresponding float not found in Excel for {xyz_filename}.")
        else:
            print(f"String '{string_to_replace}' not found in {xyz_filename}.")

if __name__ == "__main__":
    main()


